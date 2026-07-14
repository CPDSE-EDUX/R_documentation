import csv, os
from openpyxl import Workbook

DATA = os.path.dirname(os.path.abspath(__file__))  # the data/ folder

def read_csv(name, types):
    """Read a clean CSV, coercing each column with the given callables."""
    with open(os.path.join(DATA, name), newline="") as f:
        rows = list(csv.reader(f))
    header, body = rows[0], rows[1:]
    out = []
    for r in body:
        out.append([t(v) for t, v in zip(types, r)])
    return header, out

weights_h, weights = read_csv("tablet_weights.csv", [int, str, float])
diss_h,    diss    = read_csv("dissolution.csv",    [int, int, float])
titr_h,    titr    = read_csv("titration.csv",      [str, float, float])
calib_h,   calib   = read_csv("calibration.csv",    [float, float])

# ---- lab_results.xlsx : three clean sheets --------------------------------
wb = Workbook()
def add_sheet(wb, title, header, rows, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.append(header)
    for r in rows:
        ws.append(r)

# same weights, but some balance readings never made it in -> mixed NA codes.
# One column now holds numbers AND text codes, which is the whole point of na=.
missing_codes = {2: "N/A", 7: "-", 13: "ND", 19: None}  # row index -> code
weights_missing = [list(r) for r in weights]
for idx, code in missing_codes.items():
    weights_missing[idx][2] = code

add_sheet(wb, "Weights",           weights_h, weights, first=True)
add_sheet(wb, "Dissolution",       diss_h,    diss)
add_sheet(wb, "Titration",         titr_h,    titr)
add_sheet(wb, "Calibration",       calib_h,   calib)
add_sheet(wb, "Weights (missing)", weights_h, weights_missing)
wb.save(os.path.join(DATA, "lab_results.xlsx"))

# ---- lab_results_titled.xlsx : 3 title rows above the header --------------
wb2 = Workbook()
ws = wb2.active
ws.title = "Weights"
ws.append(["Ibuprofen 400 mg Tablets - Weight QC"])
ws.append(["Instrument: Mettler XPR205  |  Operator: L. Sorensen"])
ws.append(["Exported: 2026-07-14"])
ws.append(weights_h)
for r in weights:
    ws.append(r)
wb2.save(os.path.join(DATA, "lab_results_titled.xlsx"))

print("Wrote lab_results.xlsx (sheets: Weights, Dissolution, Titration, "
      "Calibration, Weights (missing))")
print("Wrote lab_results_titled.xlsx (3 title rows, header on row 4)")
